import os
import re
import logging
from datetime import datetime, timedelta, timezone
from dotenv import load_dotenv
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import warnings
import redis

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.warnings import PTBUserWarning
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    filters,
    ContextTypes,
    ConversationHandler,
)

# Suppress the PTBUserWarning for a cleaner terminal
warnings.filterwarnings('ignore', category=PTBUserWarning)

load_dotenv()
BOT_TOKEN = os.getenv("BOT_TOKEN")
REDIS_URL = os.getenv("REDIS_URL")

# Connect to Redis
redis_client = None
if REDIS_URL:
    try:
        redis_client = redis.Redis.from_url(REDIS_URL, decode_responses=True)
        redis_client.ping()
        print("Connected to Upstash Redis database successfully!")
    except Exception as e:
        print(f"Failed to connect to Redis: {e}")
        redis_client = None

# Enable logging
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.ERROR
)
logger = logging.getLogger(__name__)

# States
(
    WAITING_FOR_ADMIN_ACTION,
    WAITING_FOR_ADMIN_FILE,
    WAITING_FOR_ADMIN_DATE_SELECTION,
    WAITING_FOR_ADMIN_DATE_MANUAL,
    WAITING_FOR_ADMIN_PRICE,
    WAITING_FOR_ADMIN_DELETE,
    WAITING_FOR_USER_FILE,
    WAITING_FOR_USER_DATE_SELECTION,
) = range(8)

os.makedirs("data/temp", exist_ok=True)

# Define red fill
RED_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# --- USER FLOW (Triggered by /start or /check) ---
async def check_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    welcome_text = (
        "👋 *Welcome to the Job Report Checker Bot!*\n\n"
        "Please upload your Job Excel file to automatically check your job IDs."
    )
    await update.message.reply_text(welcome_text, parse_mode='Markdown')
    return WAITING_FOR_USER_FILE

# --- ADMIN FLOW ---
async def admin_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not redis_client:
        await update.message.reply_text("Redis Database is not configured. Please add REDIS_URL to .env")
        return ConversationHandler.END
        
    keyboard = [
        [InlineKeyboardButton("Upload Master Report", callback_data="admin_action|upload")],
        [InlineKeyboardButton("Delete Master Report", callback_data="admin_action|delete")],
        [InlineKeyboardButton("Cancel", callback_data="admin_action|cancel")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("Admin Mode: What would you like to do?", reply_markup=reply_markup)
    return WAITING_FOR_ADMIN_ACTION

async def admin_action_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    action = query.data.split("|")[1]
    
    if action == "upload":
        await query.edit_message_text(text="Please upload the Master Excel file (.xlsx)")
        return WAITING_FOR_ADMIN_FILE
    elif action == "delete":
        available_dates = redis_client.smembers("available_dates") if redis_client else []
        if not available_dates:
            await query.edit_message_text(text="No master reports available to delete in Redis.")
            return ConversationHandler.END
            
        dates = sorted(list(available_dates), reverse=True)
        keyboard = []
        for d in dates:
            keyboard.append([InlineKeyboardButton(f"Delete {d}", callback_data=f"admin_del|{d}")])
        keyboard.append([InlineKeyboardButton("Cancel", callback_data="admin_del|cancel")])
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text(text="Select a report to delete:", reply_markup=reply_markup)
        return WAITING_FOR_ADMIN_DELETE
    else:
        await query.edit_message_text(text="Action cancelled.")
        return ConversationHandler.END

async def admin_delete_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    
    date_str = query.data.split("|")[1]
    if date_str == "cancel":
        await query.edit_message_text(text="Action cancelled.")
        return ConversationHandler.END
        
    if redis_client:
        redis_client.srem("available_dates", date_str)
        redis_client.delete(f"report:{date_str}")
        redis_client.delete(f"price:{date_str}")
        await query.edit_message_text(text=f"Master report for {date_str} has been permanently deleted from Redis.")
    else:
        await query.edit_message_text(text="Redis connection lost.")
        
    return ConversationHandler.END

async def admin_file_received(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    doc = update.message.document
    if not doc.file_name.endswith('.xlsx'):
        await update.message.reply_text("Please upload an .xlsx file. Send /cancel to abort.")
        return WAITING_FOR_ADMIN_FILE
    
    file = await context.bot.get_file(doc.file_id)
    temp_path = f"data/temp/{doc.file_id}.xlsx"
    await file.download_to_drive(temp_path)
    context.user_data['admin_temp_file'] = temp_path
    
    ist_tz = timezone(timedelta(hours=5, minutes=30))
    today = datetime.now(ist_tz).strftime("%d-%m-%Y")
    yesterday = (datetime.now(ist_tz) - timedelta(days=1)).strftime("%d-%m-%Y")
    
    keyboard = [
        [InlineKeyboardButton(f"Today ({today})", callback_data=f"admin_date|{today}")],
        [InlineKeyboardButton(f"Yesterday ({yesterday})", callback_data=f"admin_date|{yesterday}")],
        [InlineKeyboardButton("Enter manually", callback_data="admin_date|manual")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text("File received. Please select the date for this report:", reply_markup=reply_markup)
    return WAITING_FOR_ADMIN_DATE_SELECTION

async def admin_date_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    
    data = query.data.split("|")[1]
    
    if data == "manual":
        await query.edit_message_text(text="Please type the date manually (Format: DD-MM-YYYY)")
        return WAITING_FOR_ADMIN_DATE_MANUAL
    else:
        context.user_data['admin_date'] = data
        await query.edit_message_text(text=f"Selected {data}.\n\nPlease enter the price per approved job (e.g. 10 or 1.5):")
        return WAITING_FOR_ADMIN_PRICE

async def admin_date_manual(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    date_str = update.message.text.strip()
    if not re.match(r"^\d{2}-\d{2}-\d{4}$", date_str):
        await update.message.reply_text("Invalid format. Please use DD-MM-YYYY. Send /cancel to abort.")
        return WAITING_FOR_ADMIN_DATE_MANUAL
    
    context.user_data['admin_date'] = date_str
    await update.message.reply_text(f"Selected {date_str}.\n\nPlease enter the price per approved job (e.g. 10 or 1.5):")
    return WAITING_FOR_ADMIN_PRICE

async def admin_price_received(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    price_str = update.message.text.strip()
    try:
        price = float(price_str)
    except ValueError:
        await update.message.reply_text("Invalid format. Please enter a valid number (e.g. 10 or 1.5):")
        return WAITING_FOR_ADMIN_PRICE
        
    date_str = context.user_data.get('admin_date')
    return await save_admin_file(update, context, date_str, price)

async def save_admin_file(update: Update, context: ContextTypes.DEFAULT_TYPE, date_str: str, price: float) -> int:
    temp_path = context.user_data.get('admin_temp_file')
    if temp_path and os.path.exists(temp_path) and redis_client:
        processing_msg = await update.message.reply_text("Extracting UIDs and uploading to Redis Database...")
        try:
            master_wb = load_workbook(temp_path, data_only=True)
            master_ws = master_wb.active
            rejected_ids = []
            
            # Start reading master file from row 7
            for row in range(7, master_ws.max_row + 1):
                val = master_ws.cell(row=row, column=1).value
                if val is not None:
                    uid = str(val).strip()
                    if uid.endswith('.0'):
                        uid = uid[:-2]
                    
                    # Skip empty cells
                    if uid and uid.lower() not in ['nan', 'none', 'null']:
                        rejected_ids.append(uid)
            
            # Save to Redis Pipeline for speed
            pipe = redis_client.pipeline()
            
            # Clear old data if any exists for this date
            pipe.delete(f"report:{date_str}")
            
            # Add new UIDs
            if rejected_ids:
                pipe.sadd(f"report:{date_str}", *rejected_ids)
                
            # Save price
            pipe.set(f"price:{date_str}", str(price))
            
            # Add to available dates
            pipe.sadd("available_dates", date_str)
            
            pipe.execute()
            
            os.remove(temp_path)
            msg = f"✅ Master report for {date_str} processed!\n💾 Saved {len(rejected_ids)} Rejected UIDs to Redis.\n💰 Price per job: {price}"
        except Exception as e:
            logger.error(f"Error parsing master file: {e}")
            msg = "❌ Error processing file. Please start over with /admin"
            
        await processing_msg.edit_text(msg)
    else:
        await update.message.reply_text("❌ Error: Temporary file lost or Redis not connected.")
        
    return ConversationHandler.END


# --- USER FLOW ---
async def user_file_received(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    doc = update.message.document
    if not doc.file_name.endswith('.xlsx'):
        await update.message.reply_text("Please upload an .xlsx file. Send /cancel to abort.")
        return WAITING_FOR_USER_FILE
    
    if not redis_client:
        await update.message.reply_text("System is down (Redis not connected).")
        return ConversationHandler.END
    
    file = await context.bot.get_file(doc.file_id)
    temp_path = f"data/temp/{doc.file_id}.xlsx"
    await file.download_to_drive(temp_path)
    context.user_data['user_temp_file'] = temp_path
    
    # Scan for available dates from Redis
    available_dates = redis_client.smembers("available_dates") if redis_client else []
    if not available_dates:
        await update.message.reply_text("No master reports available right now. Send /cancel to abort.")
        os.remove(temp_path)
        return ConversationHandler.END
        
    dates = sorted(list(available_dates), reverse=True)[:5] # top 5 recent
    
    keyboard = []
    for d in dates:
        keyboard.append([InlineKeyboardButton(d, callback_data=f"user_date|{d}")])
    keyboard.append([InlineKeyboardButton("Cancel", callback_data="user_date|cancel")])
        
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text("File received. Please select the date to check against:", reply_markup=reply_markup)
    return WAITING_FOR_USER_DATE_SELECTION

async def user_date_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    
    date_str = query.data.split("|")[1]
    
    if date_str == "cancel":
        await query.edit_message_text(text="Action cancelled.")
        user_file_path = context.user_data.get('user_temp_file')
        if user_file_path and os.path.exists(user_file_path):
            os.remove(user_file_path)
        return ConversationHandler.END
        
    await query.edit_message_text(text=f"Selected {date_str}. Processing your file against Redis, please wait...")
    
    user_file_path = context.user_data.get('user_temp_file')
    if user_file_path and os.path.exists(user_file_path):
        output_path, msg = process_user_file(user_file_path, date_str)
        
        if output_path:
            with open(output_path, 'rb') as f:
                await query.message.reply_document(document=f, caption=msg)
            os.remove(output_path)
        else:
            await query.message.reply_text(msg)
            
        if os.path.exists(user_file_path):
            os.remove(user_file_path)
    else:
        await query.message.reply_text("Error: file lost. Please try again with /check")
        
    return ConversationHandler.END

def process_user_file(user_file_path, date_str):
    if not redis_client or not redis_client.sismember("available_dates", date_str):
        return None, "No master report found for this date. Please contact Admin."
    
    # Read the price from Redis
    price_str = redis_client.get(f"price:{date_str}")
    price_per_job = float(price_str) if price_str else 0.0
    
    # Fetch all rejected IDs into a memory set for fast lookup
    rejected_ids = redis_client.smembers(f"report:{date_str}")
    
    try:
        wb = load_workbook(user_file_path)
        ws = wb.active
        
        approved_count = 0
        rejected_count = 0
        
        # Start from row 1 for the user's uploaded file
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            val = cell.value
            if val is not None:
                # Convert to string and strip float '.0' if any
                uid = str(val).strip()
                if uid.endswith('.0'):
                    uid = uid[:-2]
                
                # Skip empty cells
                if not uid or uid.lower() in ['nan', 'none', 'null']:
                    continue
                
                if uid in rejected_ids:
                    cell.fill = RED_FILL
                    rejected_count += 1
                else:
                    approved_count += 1
                    
        output_path = user_file_path.replace(".xlsx", "_checked.xlsx")
        wb.save(output_path)
        
        # Format the stats message
        earnings = approved_count * price_per_job
        msg = f"📊 Stats for {date_str}:\n\n✅ Approved Jobs: {approved_count}\n❌ Rejected Jobs: {rejected_count}\n"
        if price_per_job > 0:
            msg += f"💰 Total Earnings: {earnings}"
        else:
            msg += f"💰 Total Earnings: Price not set"
            
        return output_path, msg
        
    except Exception as e:
        logger.error(f"Error processing user file: {e}")
        return None, "Error processing your Excel file."

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text("Operation cancelled.")
    return ConversationHandler.END

def main() -> None:
    if not BOT_TOKEN:
        print("Please set BOT_TOKEN in .env")
        return

    application = Application.builder().token(BOT_TOKEN).build()

    admin_conv_handler = ConversationHandler(
        entry_points=[CommandHandler("admin", admin_start)],
        states={
            WAITING_FOR_ADMIN_ACTION: [CallbackQueryHandler(admin_action_callback, pattern=r"^admin_action\|")],
            WAITING_FOR_ADMIN_FILE: [MessageHandler(filters.Document.ALL, admin_file_received)],
            WAITING_FOR_ADMIN_DATE_SELECTION: [CallbackQueryHandler(admin_date_callback, pattern=r"^admin_date\|")],
            WAITING_FOR_ADMIN_DATE_MANUAL: [MessageHandler(filters.TEXT & ~filters.COMMAND, admin_date_manual)],
            WAITING_FOR_ADMIN_PRICE: [MessageHandler(filters.TEXT & ~filters.COMMAND, admin_price_received)],
            WAITING_FOR_ADMIN_DELETE: [CallbackQueryHandler(admin_delete_callback, pattern=r"^admin_del\|")],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        allow_reentry=True,
    )
    
    user_conv_handler = ConversationHandler(
        entry_points=[
            CommandHandler("check", check_start),
            CommandHandler("start", check_start)
        ],
        states={
            WAITING_FOR_USER_FILE: [MessageHandler(filters.Document.ALL, user_file_received)],
            WAITING_FOR_USER_DATE_SELECTION: [CallbackQueryHandler(user_date_callback, pattern=r"^user_date\|")],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        allow_reentry=True,
    )

    application.add_handler(admin_conv_handler)
    application.add_handler(user_conv_handler)

    print("Bot is starting...")
    application.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == "__main__":
    main()
